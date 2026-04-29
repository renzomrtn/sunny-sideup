import csv
import io
import asyncio
import uuid
from collections import defaultdict
from datetime import datetime

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_
from typing import Optional
import math
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from app.core.config import settings
from app.core.database import AsyncSessionLocal, get_db
from app.core.security import get_current_account, require_account_management_admin
from app.domains.identity.models import Account, AccountTenantRole, Tenant, JobPosition, LoginEvent, Outbox
from app.domains.identity.schemas import (
    AccountCreate, AccountUpdate, AccountRead, AccountListItem,
    PaginatedResponse, LoginEventRead, AccountImportResponse,
    AccountImportIssue, AccountImportWarning, AccountImportJobAccepted,
    AccountImportJobStatus,
)
from pydantic import BaseModel
from app.domains.identity.services.role_capacity import get_position_capacity
from app.services.authentik_client import (
    create_authentik_user,
    delete_authentik_user,
    resolve_authentik_user_uuid,
    set_authentik_user_active,
)

router = APIRouter(dependencies=[Depends(require_account_management_admin)])

IMPORT_JOBS: dict[str, dict] = {}

IMPORT_COLUMN_ALIASES = {
    "auth_external_id": {"auth_external_id", "auth external id", "auth id", "external id", "authentik external id"},
    "first_name":       {"first_name", "first name"},
    "middle_name":      {"middle_name", "middle name"},
    "last_name":        {"last_name", "last name"},
    "contact_number":   {"contact_number", "contact number", "phone", "mobile"},
    "contact_email":    {"contact_email", "contact email", "email"},
    "account_status":   {"account_status", "account status"},
    "tenant_name":      {"tenant_name", "tenant name", "tenant", "barangay"},
    "position_name":    {"position_name", "position name", "position", "role"},
    "is_primary_tenant":{"is_primary_tenant", "is primary", "primary tenant"},
    "role_status":      {"role_status", "role status"},
}


# ── Outbox helper ──────────────────────────────────────────────────────────

def _actor_snapshot(actor: Account) -> dict:
    primary = next(
        (r for r in actor.tenant_roles if r.is_primary_tenant and r.role_status == "Active"),
        next((r for r in actor.tenant_roles if r.role_status == "Active"), None),
    )
    return {
        "account_id":       actor.account_id,
        "auth_external_id": actor.auth_external_id,
        "full_name":        actor.full_name,
        "role_id":          primary.role_id  if primary else None,
        "tenant_id":        primary.tenant_id if primary else None,
    }


def _enqueue(db, actor: Account, event_type: str, aggregate_id: int,
             extra: dict | None = None, tenant_id: int | None = None) -> None:
    snap = _actor_snapshot(actor)
    _enqueue_snapshot(db, snap, event_type, aggregate_id, extra, tenant_id)


def _enqueue_snapshot(db, snap: dict, event_type: str, aggregate_id: int,
                      extra: dict | None = None, tenant_id: int | None = None) -> None:
    tid  = tenant_id or snap.get("tenant_id") or 0
    payload = {"actor": snap}
    if extra:
        payload.update(extra)
    db.add(Outbox(
        tenant_id=tid, aggregate_type="account",
        aggregate_id=aggregate_id, event_type=event_type, payload=payload,
    ))


# ── Misc helpers ───────────────────────────────────────────────────────────

def _normalize_text(v) -> str:
    return " ".join(str(v or "").strip().split())

def _normalize_key(v) -> str:
    return _normalize_text(v).lower()

def _normalize_name(fn, mn, ln) -> tuple[str, str, str]:
    return (_normalize_key(fn), _normalize_key(mn), _normalize_key(ln))

def _person_keys(first_name, middle_name, last_name, contact_email, contact_number) -> list[tuple[str, str]]:
    keys: list[tuple[str, str]] = []
    email = _normalize_key(contact_email)
    if email:
        keys.append(("email", email))
    name = "|".join(_normalize_name(first_name, middle_name, last_name))
    if name.strip("|"):
        number = _normalize_key(contact_number)
        if number:
            keys.append(("name_number", f"{name}|{number}"))
        keys.append(("name", name))
    return keys

def _parse_bool(v) -> bool:
    return _normalize_key(v) in {"1", "true", "yes", "y"}

def _parse_sheet_rows(filename: str, payload: bytes) -> list[dict]:
    if filename.lower().endswith(".csv"):
        return list(csv.DictReader(io.StringIO(payload.decode("utf-8-sig"))))
    if filename.lower().endswith(".xlsx"):
        wb    = load_workbook(io.BytesIO(payload), data_only=True)
        rows  = list(wb.active.iter_rows(values_only=True))
        if not rows:
            return []
        hdrs = [str(c or "").strip() for c in rows[0]]
        return [
            {hdrs[i]: row[i] if i < len(row) else None for i in range(len(hdrs))}
            for row in rows[1:]
            if any(c not in (None, "") for c in row)
        ]
    raise HTTPException(400, "Only .csv and .xlsx files are supported")

def _canonicalize_row(raw: dict) -> dict:
    out = {}
    for rk, v in raw.items():
        k = _normalize_key(rk)
        ck = next((t for t, a in IMPORT_COLUMN_ALIASES.items() if k in a), None)
        if ck:
            out[ck] = v
    return out

def _validate_role_assignment(tenant: Tenant, position: JobPosition) -> str | None:
    if tenant.is_federation and not position.is_federation_role:
        return "Federation tenants only allow federation positions"
    if not tenant.is_federation and not position.is_barangay_sk_role:
        return "Barangay tenants only allow barangay SK positions"
    return None


def _job_snapshot(job_id: str) -> dict:
    job = IMPORT_JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "Import job not found")
    return {
        "job_id": job_id,
        "status": job["status"],
        "total_rows": job["total_rows"],
        "processed_rows": job["processed_rows"],
        "message": job.get("message"),
        "result": job.get("result"),
        "error": job.get("error"),
        "created_at": job["created_at"],
        "started_at": job.get("started_at"),
        "finished_at": job.get("finished_at"),
    }


def _set_job_progress(job_id: str | None, processed_rows: int, message: str | None = None) -> None:
    if not job_id or job_id not in IMPORT_JOBS:
        return
    IMPORT_JOBS[job_id]["processed_rows"] = processed_rows
    if message:
        IMPORT_JOBS[job_id]["message"] = message


# ── List Accounts ──────────────────────────────────────────────────────────

@router.get("", response_model=PaginatedResponse)
async def list_accounts(
    page:        int           = Query(1, ge=1),
    page_size:   int           = Query(20, ge=1, le=100),
    search:      Optional[str] = Query(None),
    status:      Optional[str] = Query(None),
    tenant_id:   Optional[int] = Query(None),
    position_id: Optional[int] = Query(None),
    _:           Account       = Depends(get_current_account),
    db:          AsyncSession  = Depends(get_db),
):
    query       = select(Account)
    joined_role = False

    if search:
        term  = f"%{search}%"
        query = query.where(or_(
            Account.first_name.ilike(term),
            Account.last_name.ilike(term),
            Account.contact_email.ilike(term),
        ))
    if status:
        query = query.where(Account.account_status == status)
    if tenant_id:
        query = query.join(AccountTenantRole); joined_role = True
        query = query.where(AccountTenantRole.tenant_id == tenant_id)
    if position_id:
        if not joined_role:
            query = query.join(AccountTenantRole)
        query = query.where(AccountTenantRole.position_id == position_id)

    if joined_role:
        query = query.distinct()
    total = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar()

    # Sort: admin (account_id=1) always first, then newest → oldest
    from sqlalchemy import case as sa_case
    query = query.order_by(
        sa_case((Account.account_id == 1, 0), else_=1),
        Account.account_id.desc(),
    )
    accounts = (await db.execute(query.offset((page - 1) * page_size).limit(page_size))).scalars().all()

    items = []
    for acc in accounts:
        primary = (
            next((r for r in acc.tenant_roles if r.is_primary_tenant and r.role_status == "Active"), None)
            or next((r for r in acc.tenant_roles if r.role_status == "Active"), None)
        )
        t_name = p_name = None
        if primary:
            t = await db.get(Tenant,      primary.tenant_id)
            p = await db.get(JobPosition, primary.position_id)
            t_name = t.tenant_name   if t else None
            p_name = p.position_name if p else None
        items.append(AccountListItem(
            account_id=acc.account_id, full_name=acc.full_name,
            contact_email=acc.contact_email, account_status=acc.account_status,
            synced_at=acc.synced_at, primary_tenant=t_name, primary_position=p_name,
        ))

    return PaginatedResponse(
        items=[i.model_dump() for i in items], total=total,
        page=page, page_size=page_size,
        total_pages=math.ceil(total / page_size) if total else 1,
    )


# ── Get Single Account ─────────────────────────────────────────────────────

@router.get("/template/import")
async def download_import_template(_: Account = Depends(get_current_account), db: AsyncSession = Depends(get_db)):
    wb = Workbook(); ws = wb.active; ws.title = "Accounts Import Template"
    ws.append(["Auth External ID","First Name","Middle Name","Last Name","Contact Email",
               "Contact Number","Account Status","Tenant Name","Position Name","Is Primary Tenant","Role Status"])
    ws.append(["authentik-user-001","Juan","Santos","Dela Cruz","juan.delacruz@example.com","09171234567","Active","SK Abella","SK Chairperson","Yes","Active"])
    ws.append(["authentik-user-002","Maria","Lopez","Reyes","maria.reyes@example.com","09181234567","Active","SK Federation","Chief of Staff","No","Active"])
    notes = wb.create_sheet("Notes")
    for row in [["Field","Accepted Values / Format"],["Account Status","Active or Inactive"],
                ["Role Status","Active or Inactive"],["Is Primary Tenant","Yes/No, True/False, or 1/0"],
                ["Tenant Name","Use an exact tenant name from the system"],
                ["Position Name","Use an exact position name from the system"],
                ["Matching behavior","Accounts are matched by Auth External ID only. Same name != same person."]]:
        notes.append(row)
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": 'attachment; filename="accounts-import-template.xlsx"'})


# ── Update Account ─────────────────────────────────────────────────────────

@router.get("/{account_id}", response_model=AccountRead)
async def get_account(
    account_id: int,
    _:          Account      = Depends(get_current_account),
    db:         AsyncSession = Depends(get_db),
):
    from app.domains.identity.schemas import RoleInAccount
    account = await db.get(Account, account_id)
    if not account:
        raise HTTPException(404, "Account not found")

    has_explicit = any(r.is_primary_tenant for r in account.tenant_roles)
    first_active = next((r.role_id for r in account.tenant_roles if r.role_status == "Active"), None) \
                   if not has_explicit else None

    roles = []
    for r in account.tenant_roles:
        t = await db.get(Tenant,      r.tenant_id)
        p = await db.get(JobPosition, r.position_id)
        roles.append(RoleInAccount(
            role_id=r.role_id, tenant_id=r.tenant_id,
            tenant_name=t.tenant_name if t else "Unknown",
            position_id=r.position_id,
            position_name=p.position_name if p else "Unknown",
            is_primary_tenant=r.is_primary_tenant or (r.role_id == first_active),
            role_status=r.role_status,
        ))

    return AccountRead(
        account_id=account.account_id, auth_external_id=account.auth_external_id,
        identity_provider=account.identity_provider, first_name=account.first_name,
        middle_name=account.middle_name, last_name=account.last_name,
        contact_number=account.contact_number, contact_email=account.contact_email,
        account_status=account.account_status, synced_at=account.synced_at,
        full_name=account.full_name, tenant_roles=roles,
    )


# ── Create Account ─────────────────────────────────────────────────────────

@router.post("", response_model=AccountRead, status_code=201)
async def create_account(
    payload: AccountCreate,
    actor:   Account      = Depends(get_current_account),
    db:      AsyncSession = Depends(get_db),
):
    if (await db.execute(select(Account).where(Account.auth_external_id == payload.auth_external_id))).scalar_one_or_none():
        raise HTTPException(409, "An account with this Authentik ID already exists")

    account = Account(**payload.model_dump())
    db.add(account)
    await db.flush()
    _enqueue(db, actor, "created", account.account_id, {"account": payload.model_dump()})
    await db.commit()
    await db.refresh(account)
    return await get_account(account.account_id, actor, db)


# ── Import Accounts ────────────────────────────────────────────────────────

async def _process_account_import(
    rows: list[dict],
    actor_snapshot: dict,
    db: AsyncSession,
    job_id: str | None = None,
) -> AccountImportResponse:
    tenants   = (await db.execute(select(Tenant))).scalars().all()
    positions = (await db.execute(select(JobPosition))).scalars().all()
    accounts  = (await db.execute(select(Account))).scalars().all()
    roles_all = (await db.execute(select(AccountTenantRole))).scalars().all()

    tenant_by_name        = {t.tenant_name:   t for t in tenants}
    position_by_name      = {p.position_name: p for p in positions}
    position_by_id        = {p.position_id:   p for p in positions}
    account_by_auth_id    = {_normalize_key(a.auth_external_id): a for a in accounts if a.auth_external_id}
    account_by_person: dict[tuple[str, str], Account] = {}
    for account in accounts:
        for key in _person_keys(
            account.first_name,
            account.middle_name,
            account.last_name,
            account.contact_email,
            account.contact_number,
        ):
            account_by_person.setdefault(key, account)
    roles_by_acct_tenant  = {(r.account_id, r.tenant_id): r for r in roles_all}

    active_counts: dict[tuple[int, str], int] = defaultdict(int)
    for r in roles_all:
        if r.role_status == "Active":
            p = next((x for x in positions if x.position_id == r.position_id), None)
            if p:
                active_counts[(r.tenant_id, p.position_name)] += 1

    created = updated = roles_created = roles_updated = 0
    errors:      list[AccountImportIssue]   = []
    warnings:    list[AccountImportWarning] = []
    imported_ids: list[int] = []

    for row_index, raw in enumerate(rows, start=1):
        row_num = row_index + 1
        _set_job_progress(job_id, row_index - 1, f"Processing row {row_index} of {len(rows)}")
        row             = _canonicalize_row(raw)
        first_name      = _normalize_text(row.get("first_name"))
        middle_name     = _normalize_text(row.get("middle_name"))
        last_name       = _normalize_text(row.get("last_name"))
        auth_id         = _normalize_text(row.get("auth_external_id"))
        contact_number  = _normalize_text(row.get("contact_number")) or None
        contact_email   = _normalize_text(row.get("contact_email"))  or None
        account_status  = _normalize_text(row.get("account_status")) or "Active"
        tenant_name     = _normalize_text(row.get("tenant_name"))
        position_name   = _normalize_text(row.get("position_name"))
        role_status     = _normalize_text(row.get("role_status"))    or "Active"
        is_primary      = _parse_bool(row.get("is_primary_tenant"))

        if not first_name or not last_name:
            errors.append(AccountImportIssue(row=row_num, message="First name and last name are required")); continue
        if account_status not in {"Active", "Inactive"}:
            errors.append(AccountImportIssue(row=row_num, message="Account status must be Active or Inactive")); continue
        if role_status not in {"Active", "Inactive"}:
            errors.append(AccountImportIssue(row=row_num, message="Role status must be Active or Inactive")); continue
        if bool(tenant_name) != bool(position_name):
            errors.append(AccountImportIssue(row=row_num, message="Tenant and position must both be provided when importing roles")); continue

        person_keys = _person_keys(first_name, middle_name, last_name, contact_email, contact_number)
        matched_person_account = next(
            (account_by_person[key] for key in person_keys if key in account_by_person),
            None,
        )

        # If no auth_external_id is supplied, reuse the existing person when
        # this row is another persona; otherwise create one Authentik user.
        try:
            if auth_id:
                resolved_auth_id = await resolve_authentik_user_uuid(auth_id)
                if resolved_auth_id != auth_id:
                    warnings.append(AccountImportWarning(
                        row=row_num,
                        message="Auth External ID was converted to the Authentik UUID used for login.",
                    ))
                auth_id = resolved_auth_id
            elif matched_person_account:
                auth_id = matched_person_account.auth_external_id
            else:
                auth_id = await create_authentik_user(
                    first_name=first_name,
                    last_name=last_name,
                    email=contact_email,
                )
        except Exception as exc:
            try:
                label = "Authentik user lookup failed" if auth_id else "Authentik user creation failed"
                errors.append(AccountImportIssue(row=row_num, message=f"{label}: {exc}"))
                continue
            finally:
                _set_job_progress(job_id, row_index)

        account = account_by_auth_id.get(_normalize_key(auth_id)) or matched_person_account

        # Seed admin account is immutable — skip silently on re-import
        if account and account.auth_external_id == settings.SEED_ADMIN_AUTH_ID:
            warnings.append(AccountImportWarning(
                row=row_num,
                message="Row skipped: the System Administrator account cannot be modified via import.",
            ))
            continue

        tenant = position = None
        if tenant_name and position_name:
            tenant = tenant_by_name.get(tenant_name)
            if not tenant:
                errors.append(AccountImportIssue(row=row_num, message=f"Unknown tenant: {tenant_name}")); continue
            position = position_by_name.get(position_name)
            if not position:
                errors.append(AccountImportIssue(row=row_num, message=f"Unknown position: {position_name}")); continue
            err = _validate_role_assignment(tenant, position)
            if err:
                errors.append(AccountImportIssue(row=row_num, message=err)); continue

        role = roles_by_acct_tenant.get((account.account_id, tenant.tenant_id)) if account and tenant else None

        if tenant and position and role_status == "Active":
            cap = get_position_capacity(position.position_name)
            cnt = active_counts[(tenant.tenant_id, position.position_name)]
            if not (role and role.role_status == "Active" and role.position_id == position.position_id) and cnt >= cap:
                errors.append(AccountImportIssue(row=row_num, message=f"{position.position_name} is already full for tenant {tenant.tenant_name}")); continue

        if account:
            # Existing account — update profile fields in place
            account.first_name     = first_name
            account.middle_name    = middle_name or None
            account.last_name      = last_name
            account.contact_number = contact_number
            account.contact_email  = contact_email
            account.account_status = account_status
            updated += 1
        else:
            # New account — create it
            account = Account(
                auth_external_id=auth_id, first_name=first_name, middle_name=middle_name or None,
                last_name=last_name, contact_number=contact_number,
                contact_email=contact_email, account_status=account_status,
            )
            db.add(account)
            await db.flush()
            account_by_auth_id[_normalize_key(auth_id)] = account
            created += 1

        for key in person_keys:
            account_by_person.setdefault(key, account)

        imported_ids.append(account.account_id)

        if not tenant or not position:
            continue

        if is_primary:
            for ep in (await db.execute(select(AccountTenantRole).where(
                AccountTenantRole.account_id == account.account_id,
                AccountTenantRole.is_primary_tenant == True,
                AccountTenantRole.tenant_id != tenant.tenant_id,
            ))).scalars().all():
                ep.is_primary_tenant = False

        if role:
            op = position_by_id.get(role.position_id)
            if role.role_status == "Active" and op:
                active_counts[(tenant.tenant_id, op.position_name)] -= 1
            role.position_id = position.position_id; role.auth_external_id = account.auth_external_id
            role.is_primary_tenant = is_primary; role.role_status = role_status
            roles_updated += 1
        else:
            role = AccountTenantRole(
                account_id=account.account_id, auth_external_id=account.auth_external_id,
                tenant_id=tenant.tenant_id, position_id=position.position_id,
                is_primary_tenant=is_primary, role_status=role_status,
            )
            db.add(role); roles_by_acct_tenant[(account.account_id, tenant.tenant_id)] = role; roles_created += 1

        if role.role_status == "Active":
            active_counts[(tenant.tenant_id, position.position_name)] += 1

    if imported_ids:
        _enqueue_snapshot(db, actor_snapshot, "imported", 0, {
            "imported_account_ids": imported_ids,
            "created": created, "updated": updated,
            "roles_created": roles_created, "roles_updated": roles_updated,
        })

    await db.commit()
    _set_job_progress(job_id, len(rows), "Import complete")
    return AccountImportResponse(
        created=created, updated=updated, roles_created=roles_created,
        roles_updated=roles_updated, skipped=len(errors), errors=errors, warnings=warnings,
    )


async def _run_import_job(job_id: str, rows: list[dict], actor_snapshot: dict) -> None:
    IMPORT_JOBS[job_id].update({
        "status": "running",
        "started_at": datetime.utcnow(),
        "message": "Import started",
    })
    async with AsyncSessionLocal() as db:
        try:
            result = await _process_account_import(rows, actor_snapshot, db, job_id)
            IMPORT_JOBS[job_id].update({
                "status": "succeeded",
                "processed_rows": len(rows),
                "message": "Import complete",
                "result": result.model_dump(),
                "finished_at": datetime.utcnow(),
            })
        except Exception as exc:
            await db.rollback()
            IMPORT_JOBS[job_id].update({
                "status": "failed",
                "message": "Import failed",
                "error": str(exc),
                "finished_at": datetime.utcnow(),
            })


@router.post("/import", response_model=AccountImportJobAccepted, status_code=202)
async def import_accounts(
    file:  UploadFile   = File(...),
    actor: Account      = Depends(get_current_account),
):
    if not file.filename:
        raise HTTPException(400, "A spreadsheet file is required")
    rows = _parse_sheet_rows(file.filename, await file.read())
    if not rows:
        raise HTTPException(400, "The uploaded spreadsheet is empty")

    job_id = str(uuid.uuid4())
    IMPORT_JOBS[job_id] = {
        "status": "queued",
        "total_rows": len(rows),
        "processed_rows": 0,
        "message": "Import queued",
        "result": None,
        "error": None,
        "created_at": datetime.utcnow(),
        "started_at": None,
        "finished_at": None,
    }
    asyncio.create_task(_run_import_job(job_id, rows, _actor_snapshot(actor)))
    return AccountImportJobAccepted(
        job_id=job_id,
        status="queued",
        total_rows=len(rows),
        processed_rows=0,
        message="Import queued",
    )


@router.get("/import/{job_id}", response_model=AccountImportJobStatus)
async def get_import_job(
    job_id: str,
    _: Account = Depends(get_current_account),
):
    return _job_snapshot(job_id)


@router.patch("/{account_id}", response_model=AccountRead)
async def update_account(
    account_id: int,
    payload:    AccountUpdate,
    actor:      Account      = Depends(get_current_account),
    db:         AsyncSession = Depends(get_db),
):
    account = await db.get(Account, account_id)
    if not account:
        raise HTTPException(404, "Account not found")
    if account_id == 1:
        raise HTTPException(403, "The System Administrator account cannot be modified.")
    update_data = payload.model_dump(exclude_unset=True)
    changes = [{"field": k, "old": getattr(account, k), "new": v} for k, v in update_data.items() if getattr(account, k) != v]
    for k, v in update_data.items():
        setattr(account, k, v)
    _enqueue(db, actor, "updated", account_id, {"changes": changes})
    await db.commit(); await db.refresh(account)
    return await get_account(account_id, actor, db)


# ── Deactivate / Reactivate ────────────────────────────────────────────────

@router.patch("/{account_id}/deactivate")
async def deactivate_account(account_id: int, actor: Account = Depends(get_current_account), db: AsyncSession = Depends(get_db)):
    account = await db.get(Account, account_id)
    if not account: raise HTTPException(404, "Account not found")
    if account_id == 1: raise HTTPException(403, "The System Administrator account cannot be deactivated.")
    account.account_status = "Inactive"
    for r in account.tenant_roles: r.role_status = "Inactive"
    _enqueue(db, actor, "deactivated", account_id, {"changes": [{"field": "account_status", "old": "Active", "new": "Inactive"}]})
    await db.commit()
    try:
        await delete_authentik_user(account.auth_external_id)
    except Exception as exc:
        import logging; logging.getLogger(__name__).warning("Authentik deletion sync failed %d: %s", account_id, exc)
    return {"message": f"Account {account_id} deactivated"}


@router.patch("/{account_id}/activate")
async def activate_account(account_id: int, actor: Account = Depends(get_current_account), db: AsyncSession = Depends(get_db)):
    account = await db.get(Account, account_id)
    if not account: raise HTTPException(404, "Account not found")
    account.account_status = "Active"
    _enqueue(db, actor, "activated", account_id, {"changes": [{"field": "account_status", "old": "Inactive", "new": "Active"}]})
    await db.commit()
    try:
        await set_authentik_user_active(account.auth_external_id, True)
    except Exception as exc:
        import logging; logging.getLogger(__name__).warning("Authentik activation sync failed %d: %s", account_id, exc)
    return {"message": f"Account {account_id} activated"}


# ── Bulk Deactivate ───────────────────────────────────────────────────────

class BulkDeactivateRequest(BaseModel):
    account_ids: list[int]

@router.post("/bulk-deactivate")
async def bulk_deactivate(
    payload: BulkDeactivateRequest,
    actor:   Account      = Depends(get_current_account),
    db:      AsyncSession = Depends(get_db),
):
    account_ids = payload.account_ids
    if not account_ids:
        raise HTTPException(400, "No account IDs provided.")

    deactivated = []
    skipped = []
    for aid in account_ids:
        if aid == 1:
            skipped.append(aid)
            continue
        account = await db.get(Account, aid)
        if not account:
            skipped.append(aid)
            continue
        account.account_status = "Inactive"
        for role in account.tenant_roles:
            role.role_status = "Inactive"
        _enqueue(db, actor, "deactivated", aid, {
            "changes": [{"field": "account_status", "old": "Active", "new": "Inactive"}],
        })
        deactivated.append(aid)

    await db.commit()

    # Delete all deactivated accounts in Authentik (best-effort)
    for aid in deactivated:
        acc = await db.get(Account, aid)
        if acc:
            try:
                await delete_authentik_user(acc.auth_external_id)
            except Exception as exc:
                import logging; logging.getLogger(__name__).warning(
                    "Authentik bulk deletion sync failed %d: %s", aid, exc
                )

    return {"deactivated": deactivated, "skipped": skipped}


# ── Login History ──────────────────────────────────────────────────────────

@router.get("/{account_id}/login-history", response_model=list[LoginEventRead])
async def login_history(account_id: int, limit: int = Query(50, ge=1, le=200), _: Account = Depends(get_current_account), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(LoginEvent).where(LoginEvent.account_id == account_id).order_by(LoginEvent.occurred_at.desc()).limit(limit))
    return result.scalars().all()
